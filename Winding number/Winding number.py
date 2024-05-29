"""
This code is supported by the website: https://www.guanjihuan.com
The newest version of this code is on the web page: https://www.guanjihuan.com/archives/5025
"""
import xlsxwriter as xw
import numpy as np
import matplotlib.pyplot as plt
from math import *
import cmath
import time
import random
import pandas as pd
import openpyxl as op
bg = op.load_workbook("Test.xlsx")      	# 应先将excel文件放入到工作目录下
sheet = bg["Sheet1"]
arr1=[0]
arr2=[0]
arr3=[0]
fileName = "Test.xls" #工作簿名字
workbook = xw.Workbook(fileName)  #创建工作簿
worksheet1 = workbook.add_worksheet("sheet1")  # 创建子表
worksheet1.activate()  # 激活表
count=0
# ws.create_sheet('sheet2') #创建一个名为sheet2的作薄(一般不用写默认会自动创建)

def hamiltonian(k,r1,r2):  # SSH模型

    v = r1
    w = r2
    matrix = np.zeros((2, 2), dtype=complex)
    matrix[0, 1] = v + w * cmath.exp(-1j * k)
    matrix[1, 0] = v + w * cmath.exp(1j * k)
    return matrix


for i in range(10000):
    count=count+1
    print(count)
    start_clock = time.perf_counter()
    delta_1 = 1e-9  # 求导的步长（求导的步长可以尽可能短）
    delta_2 = 1e-5  # 积分的步长（积分步长和计算时间相关，因此取一个合理值即可）
    W = 0  # Winding number初始化
    r1 = random.random() * 4 - 2    #纵轴
    arr1.append(r1)
    r2 = random.random() * 4 - 2    #横轴
    arr2.append(r2)


    for k in np.arange(-pi, pi, delta_2):
        H = hamiltonian(k,r1,r2)
        log0 = cmath.log(H[0, 1])
        H_delta = hamiltonian(k + delta_1,r1,r2)
        log1 = cmath.log(H_delta[0, 1])
        W = W + (log1 - log0) / delta_1 * delta_2  # Winding number
    print('v=',r1,'w=',r2,W / 2 / pi / 1j)
    arr3.append(W / 2 / pi / 1j)
    arr=[[r1,r2,W / 2 / pi / 1j]]


    print('Winding number = ', W / 2 / pi / 1j)
    end_clock = time.perf_counter()
    print('CPU执行时间(min)=', (end_clock - start_clock) / 60)
    # 设置横轴的上下限值
    plt.xlim(-2, 2)
    # 设置纵轴的上下限值
    plt.ylim(-2, 2)
    if abs(W / 2 / pi /1j) < 0.1:
        plt.scatter(r2,r1,marker='o',color='red',s=9)
    else:
        plt.scatter(r2, r1, marker='o', color='blue', s=9)


plt.show()



